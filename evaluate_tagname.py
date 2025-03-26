from Evaluate.my_metrics import similarity_result_two_folders
import pandas as pd
import time
from Config import Data_num, Output_num, Type, Label_Input_num

from collections import Counter

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
import ast

# === Folder Addresses (All path should be here) ===
# label_folder = f"Temp\Data_{Data_num}\{Type}\Label{Label_Input_num}\Differents"
label_folder = f"Temp\Data_{Data_num}\{Type}\Label{Label_Input_num}\Processed_Label\Differents"
llm_filled_folder = f"Temp\Data_{Data_num}\{Type}\Output{Output_num}\Processed_Output\Differents"
root_folder = f"Temp\Data_{Data_num}\{Type}"

#ensure root_folder/Resulst folder exist
import os
if not os.path.exists(f"{root_folder}/Results"):
    os.makedirs(f"{root_folder}/Results")


# Function to process df_detail
def analyze_errors_type_1(error_list):
    """
    Phân tích danh sách lỗi, đếm số lần xuất hiện của từng lỗi (tag1, tag2).
    :param error_list: Danh sách chứa tuple (tag1, tag2)
    :return: Danh sách chứa tuple (tag1, tag2, count) sắp xếp theo số lần xuất hiện giảm dần.
    """
    counter = Counter(error_list)
    return sorted([(tag1, tag2, count) for (tag1, tag2), count in counter.items()], key=lambda x: x[2], reverse=True)

def analyze_errors_type_2(error_list):
    """
    Phân tích danh sách lỗi, đếm số lần xuất hiện của từng lỗi (tag1, tag2, context).
    :param error_list: Danh sách chứa tuple (tag1, tag2, context)
    :return: Danh sách chứa tuple (tag1, tag2, context, count) sắp xếp theo số lần xuất hiện giảm dần.
    """
    counter = Counter(error_list)
    return sorted([(tag1, tag2, context, count) for (tag1, tag2, context), count in counter.items()], 
                  key=lambda x: x[3], reverse=True)

# ============= Evaluate between label and llm_filled folders=============
df, df_detail = similarity_result_two_folders(label_folder, llm_filled_folder)
# Save to csv
time_now = time.strftime('%Y-%m-%d-%H-%M-%S')

def process_df_detail(df_detail):
    # Process df detail
    mean_columns = ["P_A1_A1", "P_B_B"]
    sum_columns = ["P_A1_A2", "P_A1_B", "P_B_A1"]
    mean_row = df_detail[mean_columns].mean()
    sum_row = df_detail[sum_columns].sum()
    # Combine mean and sum into a single DataFrame row
    summary_row = pd.concat([mean_row, sum_row], axis=0)
    summary_row["total_A1"] = "Summary"
    summary_row["C_A1_A2"] = analyze_errors_type_1(sum(df["error A1-A2"], []))
    summary_row["C_A1_B"] = analyze_errors_type_1(sum(df["error A1-B"], []))
    summary_row["C_B_A1"] = analyze_errors_type_1(sum(df["error B-A1"], []))
    summary_row["D_A1_A2"] = analyze_errors_type_2(sum(df["error A1-A2 detail"], []))
    summary_row["D_A1_B"] = analyze_errors_type_2(sum(df["error A1-B detail"], []))
    summary_row["D_B_A1"] = analyze_errors_type_2(sum(df["error B-A1 detail"], []))
    # concat
    df_detail = pd.concat([df_detail, pd.DataFrame([summary_row])], ignore_index=True)
    
    return df_detail

df_detail = process_df_detail(df_detail)
# Add user_X_X, user_X_Y column
df_detail["user_X_X"] = df["user_X_X"]
df_detail["user_X_Y"] = df["user_X_Y"]
df_detail["debug_user_X_Y_label"] = df["debug_user_X_Y_label"]
df_detail["debug_user_X_Y_predict"] = df["debug_user_X_Y_predict"]
# Save detail to csv
# df_detail.to_csv(f"{root_folder}/Result_{Output_num}_{time_now}.csv", index=False,encoding='utf-8-sig')
df_detail.to_csv(f"{root_folder}/Results/Result_{Output_num}.csv", index=False,encoding='utf-8-sig')


# Summary data
def frac(a,b):
    temp = a/b*100
    if int(temp)==temp:
        return int(temp)
    else:
        return round(temp,2)
    
def df_detail_to_summary_excel(df_detail, excel_filename, name_sheet):
    # Number of  forms
    num_forms = df_detail.shape[0]-1
    # Value
    total_A1 = df_detail["total_A1"].iloc[:-1].astype(int).sum()
    total_B = df_detail["total_B"].iloc[:-1].astype(int).sum()
    real_tagnames = total_A1 + total_B
    # Seen
    real_seen_tagnames = total_A1
    real_seen_tagnames_percent = frac(real_seen_tagnames,real_tagnames)
    # Unseen
    real_unseen_tagnames = real_tagnames - real_seen_tagnames
    real_unseen_tagnames_percent = frac(real_unseen_tagnames,real_tagnames)
    # Predicted value
    true_tagname = df_detail["P_A1_A1"].iloc[:-1].astype(int).sum()
    false_tagname_tagname = df_detail["P_A1_A2"].iloc[:-1].astype(int).sum()
    false_unseen_tagname = df_detail["P_A1_B"].iloc[:-1].astype(int).sum()
    false_tagname_unseen_tagname = df_detail["P_B_A1"].iloc[:-1].astype(int).sum()
    true_unseen_tagname = df_detail["P_B_B"].iloc[:-1].astype(int).sum()

    # Predicted percentage value - real seen tagnames
    true_tagname_percent = frac(true_tagname,real_seen_tagnames)
    false_tagname_tagname_percent = frac(false_tagname_tagname,real_seen_tagnames)
    false_unseen_tagname_percent = frac(false_unseen_tagname,real_seen_tagnames)

    # Predicted percentage value - real unseen tagnames
    false_tagname_unseen_tagname_percent = frac(false_tagname_unseen_tagname,real_unseen_tagnames)
    true_unseen_tagname_percent = frac(true_unseen_tagname,real_unseen_tagnames)

    # Process
    # Create a 5x5 DataFrame (initial template)
    data = [[f"R{r}C{c}" for c in range(1, 6)] for r in range(1, 6)]
    df = pd.DataFrame(data)

    # Save to Excel first
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=name_sheet, index=False, header=False)
    # Load workbook and worksheet
    wb = load_workbook(excel_filename)
    ws = wb[name_sheet]  # Access the name_sheet sheet  

    # Merge A1:C2 (Excel: A1:C2)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=3)
    ws["A1"].value = f"Number of forms \n {num_forms}"  
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge A3:A5 and set value to "Predicted \nTagnames"
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=1)
    ws["A3"].value = "Predicted \nTagnames"
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge D1:E1 and set value to f"Real Tagnames ({real_tagnames})"
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    ws["D1"].value = f"Real Tagnames ({real_tagnames})"
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set D2 to "Seen Tagnames (real_seen_tagnames)" and wrap text with percentage on a new line
    ws["D2"].value = f"Seen Tagnames ({real_seen_tagnames})  ({real_seen_tagnames_percent}%)"
    ws["D2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set E2 to "Unseen Tagnames (real_unseen_tagnames)" and wrap text with percentage on a new line
    ws["E2"].value = f"Unseen Tagnames ({real_unseen_tagnames}) ({real_unseen_tagnames_percent}%)"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge B3:B4 and set value to "Seen tagname"
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws["B3"].value = "Seen tagname"
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    # Set C3 to "TRUE" and C4 to "FALSE"
    ws["C3"].value = "TRUE"
    ws["C4"].value = "FALSE"
    ws["C3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C4"].alignment = Alignment(horizontal="center", vertical="center")

    # Merge B5:C5 and set value to "Unseen tagname"
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
    ws["B5"].value = "Unseen tagname"
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")

    # Set D3, D4, D5 with the new values, including line breaks
    ws["D3"].value = f"{true_tagname}\n({true_tagname_percent}%)"
    ws["D4"].value = f"{false_tagname_tagname}\n({false_tagname_tagname_percent}%)"
    ws["D5"].value = f"{false_unseen_tagname}\n({false_unseen_tagname_percent}%)"
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Merge E3:E4 and set value to f"{false_tagname_unseen_tagname}\n({false_tagname_unseen_tagname_percent}%)"
    ws.merge_cells(start_row=3, start_column=5, end_row=4, end_column=5)
    ws["E3"].value = f"{false_tagname_unseen_tagname}\n({false_tagname_unseen_tagname_percent}%)"
    ws["E3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set E5 to f"{true_unseen_tagname}\n({true_unseen_tagname_percent}%)"
    ws["E5"].value = f"{true_unseen_tagname}\n({true_unseen_tagname_percent}%)"
    ws["E5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply borders to all cells in the range A1:E5
    border = Border(
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin")
    )

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
        for cell in row:
            cell.border = border

    # Increase row heights by 2 times
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 45  # Default is 15, so we double it

    # Increase column widths by 2 times
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20  # Default is 8.43, so we double it

    # Increase font size to 14 for all cells in A1:E5
    font_14 = Font(size=14)
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
        for cell in row:
            cell.font = font_14

    # Save the modified workbook
    wb.save(excel_filename)
    wb.close()

    print(f"Excel file '{excel_filename}' saved successfully!")

# Save Summary to Excel
# name_xlsx = f"{root_folder}/Summary_{Output_num}_{time_now}.xlsx"
name_xlsx = f"{root_folder}/Results/Summary_{Output_num}.xlsx"
name_sheet = "Summary"
df_detail_to_summary_excel (df_detail, name_xlsx, name_sheet)

# Add detail error to this summary file
def add_error_to_summary_file(df_detail, err_name, excel_filename, name_sheet):
    # Type count first
    err_count = df_detail.iloc[-1,:][f"C_{err_name}"]
    # Convert data into a list of tuples for DataFrame
    formatted_data = [(item[0], item[1], item[2]) for item in err_count]
    # Create the DataFrame
    df_err_count = pd.DataFrame(formatted_data, columns=['label', 'predict', 'num_error'])
    # Type Detail
    err_detail = df_detail.iloc[-1,:][f"D_{err_name}"]
    # Convert data into a list of tuples for DataFrame
    formatted_data = [(item[0], item[1], item[2], item[3]) for item in err_detail]
    # Create the DataFrame
    df_err_detail = pd.DataFrame(formatted_data, columns=['label_context','label', 'predict', 'num_error'])
    # Save both DataFrames into the same sheet
    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode="a",if_sheet_exists='overlay') as writer:
        df_err_count.to_excel(writer, sheet_name=name_sheet, startrow=0, index=False)
        df_err_detail.to_excel(writer, sheet_name=name_sheet, startrow=len(df_err_count)+2, index=False)
        # Access the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[name_sheet]

        # Increase row height (double)
        default_row_height = 15  # Default row height
        for row in range(1, len(df_err_count) + len(df_err_detail) + 3):  # Including gap
            worksheet.row_dimensions[row].height = default_row_height * 2  # Doubling row height

        # Increase column width (double)
        default_col_width = 8.43  # Default column width
        double_size = 4
        for col_num in range(1, len(df_err_count.columns) + 1):  # Columns of df_err_count
            worksheet.column_dimensions[chr(64 + col_num)].width = default_col_width * double_size  # Doubling column width

        for col_num in range(1, len(df_err_detail.columns) + 1):  # Columns of df_err_detail
            worksheet.column_dimensions[chr(64 + col_num + len(df_err_count.columns))].width = default_col_width * double_size  # Doubling column width

        # Apply text alignment and font size for all cells
        for row in worksheet.iter_rows(min_row=1, max_row=len(df_err_count) + len(df_err_detail) + 2, min_col=1, max_col=max(len(df_err_count.columns), len(df_err_detail.columns))):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(size=14, bold=True)

        # Save the modified workbook
        workbook.save(excel_filename)
error_A1_A2 = "A1_A2"
error_A1_B = "A1_B"
error_B_A1 = "B_A1"

# Add error_A1_A2 to the summary file
add_error_to_summary_file(df_detail, error_A1_A2, name_xlsx, name_sheet=error_A1_A2)
add_error_to_summary_file(df_detail, error_A1_B, name_xlsx, name_sheet=error_A1_B)
add_error_to_summary_file(df_detail, error_B_A1, name_xlsx, name_sheet=error_B_A1)

# Return another csv include
# Data_Num, num_forms, sum_tagname, total_a1, a1_a1, total_b, b_b, a1_a1, b_b, a1_a1, a1_b, b_a1
# Data_Num, num_forms, %sum_tagname, %total_a1, %a1_a1, %total_b, %b_b, %a1_a1, %b_b, %a1_a1, %a1_b, %b_a1
# Number of  forms
num_forms = df_detail.shape[0]-1
# Value
total_A1 = df_detail["total_A1"].iloc[:-1].astype(int).sum()
total_B = df_detail["total_B"].iloc[:-1].astype(int).sum()
real_tagnames = total_A1 + total_B
# Seen
real_seen_tagnames = total_A1
real_seen_tagnames_percent = frac(real_seen_tagnames,real_tagnames)
# Unseen
real_unseen_tagnames = real_tagnames - real_seen_tagnames
real_unseen_tagnames_percent = frac(real_unseen_tagnames,real_tagnames)
# Predicted value
true_tagname = df_detail["P_A1_A1"].iloc[:-1].astype(int).sum()
false_tagname_tagname = df_detail["P_A1_A2"].iloc[:-1].astype(int).sum()
false_unseen_tagname = df_detail["P_A1_B"].iloc[:-1].astype(int).sum()
false_tagname_unseen_tagname = df_detail["P_B_A1"].iloc[:-1].astype(int).sum()
true_unseen_tagname = df_detail["P_B_B"].iloc[:-1].astype(int).sum()

# Predicted percentage value - real seen tagnames
true_tagname_percent = frac(true_tagname,real_seen_tagnames)
false_tagname_tagname_percent = frac(false_tagname_tagname,real_seen_tagnames)
false_unseen_tagname_percent = frac(false_unseen_tagname,real_seen_tagnames)

# Predicted percentage value - real unseen tagnames
false_tagname_unseen_tagname_percent = frac(false_tagname_unseen_tagname,real_unseen_tagnames)
true_unseen_tagname_percent = frac(true_unseen_tagname,real_unseen_tagnames)

# Info about error X_Y
num_form_error_X_Y = int(sum(df_detail["user_X_Y"]!=0)-1)
detail_form_error_X_Y = df_detail.loc[df_detail["user_X_Y"] != 0, "user_X_Y"][:-1].astype("int").tolist()

# fist_row = [f"Data_{Label_Input_num}", num_forms, f"{real_tagnames}", f"{total_A1} ({frac(total_A1,real_tagnames)})", f"{true_tagname} ({frac(true_tagname,real_seen_tagnames)})", f"{total_B} ({frac(total_B,real_tagnames)})", f"{true_unseen_tagname} ({frac(true_unseen_tagname,real_unseen_tagnames)})", f"{false_tagname_tagname} ({frac(false_tagname_tagname,real_seen_tagnames)})", f"{false_unseen_tagname} ({frac(false_unseen_tagname,real_seen_tagnames)})", f"{false_tagname_unseen_tagname} ({frac(false_tagname_unseen_tagname,real_unseen_tagnames)})"]
fist_row = [f"Data_{Label_Input_num}", num_forms, f"{real_tagnames}", f"{total_A1} ({frac(total_A1,real_tagnames)}%)", f"{false_unseen_tagname} ({frac(false_unseen_tagname,real_seen_tagnames)}%)", f"{false_tagname_tagname} ({frac(false_tagname_tagname,real_seen_tagnames)}%)",f"{false_tagname_unseen_tagname} ({frac(false_tagname_unseen_tagname,real_unseen_tagnames)}%)"]
# second_row = [f"Data_{Label_Input_num}", num_forms, frac(real_tagnames,real_tagnames), frac(total_A1,real_tagnames), frac(true_tagname,real_seen_tagnames), frac(total_B,real_tagnames), frac(true_unseen_tagname,real_unseen_tagnames), frac(false_tagname_tagname,real_seen_tagnames), frac(false_unseen_tagname,real_seen_tagnames), frac(false_tagname_unseen_tagname,real_unseen_tagnames)]
third_row = [f"Data_{Label_Input_num}",num_forms, f"{num_form_error_X_Y} ({frac(num_form_error_X_Y,num_forms)}%)", detail_form_error_X_Y]
# Save to statis_{Label_Input_num}.csv
statis_csv = f"{root_folder}/Results/Result_statis_{Label_Input_num}.csv"
column_names = ["Loại dữ liệu", "Số forms", "Tổng tagname", "Seen Tagname","FUT", "FT-T", "FT-UT"]
df_statis = pd.DataFrame([fist_row, third_row], columns=column_names)
# Replace values in the "Loại dữ liệu" column
df_statis["Loại dữ liệu"] = df_statis["Loại dữ liệu"].replace({
    "Data_31": "Thực tế",
    "Data_21": "LLM",
    "Data_11": "Quy tắc"
})
df_statis.to_csv(statis_csv, index=False,encoding='utf-8-sig')
# Print
print("Save result successfully!!")




