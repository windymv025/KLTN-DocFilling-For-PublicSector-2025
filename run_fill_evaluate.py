import re
import subprocess
import pandas as pd
# Define the path to config.py
from Config import Data_num, Type
import openpyxl

config_file = "Config/config.py"
fill_tagname_file = "fill_tagname_to_input.py"
root_folder = f"Temp\Data_{Data_num}\{Type}"
name_result = "A"


# Define the test cases for output_label_input_num and Data_num
output_label_input_nums = [31]

# Run with process_tagname = False in fill_tagname_to_input.py
for output_label_input_num in output_label_input_nums:
    print(f"\nTesting with output_label_input_num={output_label_input_num}")

    # Read and modify config.py content
    with open(config_file, "r", encoding="utf-8") as file:
        content = file.read()
    # Modify the variables
    content = re.sub(r'output_label_input_num\s*=\s*\d+', f'output_label_input_num = {output_label_input_num}', content)
    # Write back the modified content
    with open(config_file, "w", encoding="utf-8") as file:
        file.write(content)
    print("Updated config.py ✅")
    
    # Read and modify fill_tagname_to_input.py (process_tagname=False) content
    with open(fill_tagname_file, "r", encoding="utf-8") as file:
        content = file.read()
    # Modify the variables
    content = re.sub(r'process_tagname\s*=\s*\w+', 'process_tagname = False', content)
    # Write back the modified content  
    with open(fill_tagname_file, "w", encoding="utf-8") as file:
        file.write(content)
    print("Updated fill_tagname_to_input.py ✅")

    # Run fill.py
    subprocess.run(["python", "fill_tagname_to_input.py"])
    print("Executed fill.py ✅")

# Run with process_tagname = True in fill_tagname_to_input.py
for output_label_input_num in output_label_input_nums:
    print(f"\nTesting with output_label_input_num={output_label_input_num}")

    # Read config.py content
    with open(config_file, "r", encoding="utf-8") as file:
        content = file.read()
    # Modify the variables
    content = re.sub(r'output_label_input_num\s*=\s*\d+', f'output_label_input_num = {output_label_input_num}', content)
    # Write back the modified content
    with open(config_file, "w", encoding="utf-8") as file:
        file.write(content)
    print("Updated config.py ✅")

    # Read and modify fill_tagname_to_input.py (process_tagname=True) content
    with open(fill_tagname_file, "r", encoding="utf-8") as file:
        content = file.read()
    # Modify the variables
    content = re.sub(r'process_tagname\s*=\s*\w+', 'process_tagname = True', content)
    # Write back the modified content  
    with open(fill_tagname_file, "w", encoding="utf-8") as file:
        file.write(content)
    print("Updated fill_tagname_to_input.py ✅")

    # Run fill.py
    subprocess.run(["python", "fill_tagname_to_input.py"])
    print("Executed fill.py ✅")

# Run evaluate
for output_label_input_num in output_label_input_nums:
    print(f"\nTesting with output_label_input_num={output_label_input_num}")
    # Read config.py content
    with open(config_file, "r", encoding="utf-8") as file:
        content = file.read()
    # Modify the variables
    content = re.sub(r'output_label_input_num\s*=\s*\d+', f'output_label_input_num = {output_label_input_num}', content)
    # Write back the modified content
    with open(config_file, "w", encoding="utf-8") as file:
        file.write(content)
    print("Updated config.py ✅")
    # Run evaluate.py
    subprocess.run(["python", "evaluate_tagname.py"])
    print("Executed evaluate.py ✅")
    

sum_df = pd.read_csv(f"{root_folder}/Results/Result_statis_{output_label_input_nums[0]}.csv")
X_Y_error = sum_df.iloc[1:,:]
sum_df = sum_df.iloc[:1,:]
for output_label_input_num in output_label_input_nums[1:]:
    df = pd.read_csv(f"{root_folder}/Results/Result_statis_{output_label_input_num}.csv")
    X_Y_error_Df = df.iloc[1:,:]
    df = df.iloc[:1,:]
    sum_df = pd.concat([sum_df, df], ignore_index=True)
    X_Y_error = pd.concat([X_Y_error, X_Y_error_Df], ignore_index=True)

# df_reordered = pd.concat([sum_df.iloc[0:-1:2], sum_df.iloc[1:-1:2]]).reset_index(drop=True)


sum_df.to_csv(f"{root_folder}/{name_result}_result_fieldname.csv", index=False, encoding="utf-8-sig")

X_Y_error = X_Y_error.iloc[:,:3]
X_Y_error.columns = ["Loại dữ liệu","Số forms","Lỗi sai"]
X_Y_error.to_csv(f"{root_folder}/{name_result}_result_userX.csv", index=False, encoding="utf-8-sig")
## Merge all summary to one --> debugging
list_summary_file = [f"{root_folder}/Results/Summary_{output_label_input_num}.xlsx" for output_label_input_num in output_label_input_nums]
sheets_to_merge = ["A1_A2", "A1_B", "B_A1"]  # Sheets to merge
output_merge_summary_file = f"{root_folder}/{name_result}_Merged_Summary.xlsx"
# Create a new workbook
merged_wb = openpyxl.Workbook()
merged_wb.remove(merged_wb.active)  # Remove default sheet

for index, file in enumerate(list_summary_file):
    wb = openpyxl.load_workbook(file, data_only=True)

    for sheet_name in sheets_to_merge:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Create a new sheet in merged workbook with custom name
            new_sheet_name = f"{output_label_input_nums[index]}_{sheet_name}"
            new_ws = merged_wb.create_sheet(title=new_sheet_name)

            # Copy content to new sheet
            for row in ws.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate].value = cell.value

            # Copy merged cells
            for merged_range in ws.merged_cells.ranges:
                new_ws.merge_cells(str(merged_range))

# Save the merged workbook
merged_wb.save(output_merge_summary_file)
print("Merged all summary files to one ✅")